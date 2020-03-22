from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from PIL import Image


class ExcelPixelator:
    def __init__(self, input_path, output_path, file_name, cell_size, pixel_size):
        self.image = Image.open(input_path).convert('RGB')
        self.output_path = output_path
        self.file_name = file_name
        self.cell_size = cell_size  # size of the cell in pixels
        self.pixel_size = pixel_size

        self.__is_pixelsize_common_factor()

    def create_pixel_art(self):
        default_excel_font_size = 16
        wb = Workbook()
        ws = wb.create_sheet('Pixel-Art')
        wb.remove(wb['Sheet'])  # remove default worksheet

        width, height = self.image.size
        pixel_map = self.get_pixel_map(width, height)

        for row in range(len(pixel_map)):
            ws.row_dimensions[row + 1].height = self.cell_size * 10 / default_excel_font_size
            for col in range(len(pixel_map[row])):
                curr_col = get_column_letter(col + 1)
                ws.column_dimensions[curr_col].width = self.cell_size / 9

                rgbTuple = pixel_map[row][col]
                fill_color = self.__rgbToHex(rgbTuple)
                ws[f'{curr_col}{row + 1}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        wb.save(f'{self.output_path}/{self.file_name}.xlsx')
        wb.close()

    def get_pixel_map(self, w, h):
        pixel_map = [[0 for x in range(w // self.pixel_size)] for y in range(h // self.pixel_size)]
        squares = w * h // self.pixel_size ** 2
        i, j = 0, 0

        for sq in range(squares):
            rAvg, gAvg, bAvg = 0, 0, 0
            row_start = (sq * self.pixel_size // h) * self.pixel_size
            row_end = row_start + self.pixel_size
            col_start = sq * self.pixel_size % w
            col_end = col_start + self.pixel_size

            for row in range(row_start, row_end):
                for col in range(col_start, col_end):
                    r, g, b = self.image.getpixel((col, row))
                    rAvg += r
                    gAvg += g
                    bAvg += b
            rAvg //= self.pixel_size ** 2
            gAvg //= self.pixel_size ** 2
            bAvg //= self.pixel_size ** 2
            pixel_map[i][j] = (rAvg, gAvg, bAvg)
            i = i + 1 if j >= (w // self.pixel_size) - 1 else i
            j = (j + 1) % (w // self.pixel_size)

        return pixel_map

    def __rgbToHex(self, rgbTuple):
        return ('%02x%02x%02x' % rgbTuple).upper()

    def __is_pixelsize_common_factor(self):
        width, height = self.image.size
        if width % self.pixel_size != 0 or height % self.pixel_size != 0:
            print('ERROR: Pixel size must be a number divisible exactly by both the image width and height')
            exit(1)
